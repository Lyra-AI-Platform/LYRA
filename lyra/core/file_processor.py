"""

Copyright (C) 2026 Lyra Contributors
Licensed under the Lyra Community License v1.0. See LICENSE for details.
Lyra File Processor
Handles uploaded files: PDFs, images, code, text, CSV, etc.
Extracts content and prepares it for AI analysis.
"""
import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional, Tuple

logger = logging.getLogger(__name__)

UPLOADS_DIR = Path(__file__).parent.parent.parent / "data" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# Max content to send to AI (chars)
MAX_CONTENT_CHARS = 50_000


class FileProcessor:
    """Process uploaded files into AI-readable content."""

    SUPPORTED_TYPES = {
        # Text/code
        ".txt": "text", ".md": "text", ".rst": "text",
        ".py": "code", ".js": "code", ".ts": "code", ".jsx": "code",
        ".tsx": "code", ".java": "code", ".c": "code", ".cpp": "code",
        ".go": "code", ".rs": "code", ".rb": "code", ".php": "code",
        ".html": "code", ".css": "code", ".sql": "code", ".sh": "code",
        ".yaml": "code", ".yml": "code", ".toml": "code", ".json": "code",
        ".xml": "code", ".env": "code",
        # Documents
        ".pdf": "pdf",
        ".docx": "docx",
        # Data
        ".csv": "csv",
        ".xlsx": "xlsx",
        # Images
        ".png": "image", ".jpg": "image", ".jpeg": "image",
        ".gif": "image", ".webp": "image", ".bmp": "image",
    }

    async def process(self, file_path: str, filename: str) -> Dict[str, Any]:
        """
        Process a file and return extracted content + metadata.
        Returns: {success, content, type, filename, size, error}
        """
        path = Path(file_path)
        if not path.exists():
            return {"success": False, "error": "File not found"}

        suffix = path.suffix.lower()
        file_type = self.SUPPORTED_TYPES.get(suffix, "unknown")
        size = path.stat().st_size

        result = {
            "success": True,
            "filename": filename,
            "type": file_type,
            "size_bytes": size,
            "size_human": self._human_size(size),
            "content": "",
            "error": None,
        }

        try:
            if file_type in ("text", "code"):
                result["content"] = self._read_text(path)
            elif file_type == "pdf":
                result["content"] = self._read_pdf(path)
            elif file_type == "docx":
                result["content"] = self._read_docx(path)
            elif file_type == "csv":
                result["content"] = self._read_csv(path)
            elif file_type == "xlsx":
                result["content"] = self._read_xlsx(path)
            elif file_type == "image":
                result["content"] = self._describe_image_prompt(path, filename)
                result["is_image"] = True
                result["image_path"] = str(path)
            else:
                # Try reading as text anyway
                try:
                    result["content"] = self._read_text(path)
                except Exception:
                    result["success"] = False
                    result["error"] = f"Unsupported file type: {suffix}"
        except Exception as e:
            result["success"] = False
            result["error"] = str(e)
            logger.error(f"File processing error ({filename}): {e}")

        return result

    def _read_text(self, path: Path) -> str:
        """Read text/code files."""
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(MAX_CONTENT_CHARS)
        if len(content) == MAX_CONTENT_CHARS:
            content += "\n\n[...file truncated for length...]"
        return content

    def _read_pdf(self, path: Path) -> str:
        """Extract text from PDF."""
        try:
            import pymupdf  # PyMuPDF (fitz)
            doc = pymupdf.open(str(path))
            pages = []
            for i, page in enumerate(doc):
                text = page.get_text()
                if text.strip():
                    pages.append(f"--- Page {i+1} ---\n{text}")
                if len("\n".join(pages)) > MAX_CONTENT_CHARS:
                    pages.append("\n[...PDF truncated...]")
                    break
            return "\n\n".join(pages)
        except ImportError:
            # Fallback to pypdf
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(path))
                pages = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text:
                        pages.append(f"--- Page {i+1} ---\n{text}")
                return "\n\n".join(pages)[:MAX_CONTENT_CHARS]
            except ImportError:
                return "[PDF reading requires: pip install pymupdf OR pip install pypdf]"

    def _read_docx(self, path: Path) -> str:
        """Extract text from Word document."""
        try:
            import docx
            doc = docx.Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)[:MAX_CONTENT_CHARS]
        except ImportError:
            return "[DOCX reading requires: pip install python-docx]"

    def _read_csv(self, path: Path) -> str:
        """Read CSV and format as readable table."""
        try:
            import csv
            rows = []
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    rows.append(" | ".join(row))
                    if i > 500:
                        rows.append(f"[...{i} more rows truncated...]")
                        break
            return "\n".join(rows)
        except Exception as e:
            return f"[CSV read error: {e}]"

    def _read_xlsx(self, path: Path) -> str:
        """Read Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            result = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                result.append(f"=== Sheet: {sheet_name} ===")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    result.append(" | ".join(str(c) if c is not None else "" for c in row))
                    if i > 500:
                        result.append("[...truncated...]")
                        break
            return "\n".join(result)[:MAX_CONTENT_CHARS]
        except ImportError:
            return "[Excel reading requires: pip install openpyxl]"

    def _describe_image_prompt(self, path: Path, filename: str) -> str:
        """
        For images, return a prompt noting the image is attached.
        Actual vision processing handled by multimodal models.
        """
        return f"[Image file attached: {filename} ({self._human_size(path.stat().st_size)})]"

    def _human_size(self, size_bytes: int) -> str:
        for unit in ["B", "KB", "MB", "GB"]:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"

    def format_for_prompt(self, processed: Dict) -> str:
        """Format processed file content as AI prompt context."""
        filename = processed.get("filename", "file")
        ftype = processed.get("type", "file")
        content = processed.get("content", "")
        size = processed.get("size_human", "")

        header = f"[ATTACHED FILE: {filename} | Type: {ftype} | Size: {size}]\n"
        if ftype == "code":
            ext = Path(filename).suffix.lstrip(".")
            return header + f"```{ext}\n{content}\n```"
        return header + content


# Global singleton
file_processor = FileProcessor()
